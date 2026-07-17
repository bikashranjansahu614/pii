import io
import openpyxl
from fastapi import FastAPI, UploadFile, HTTPException, status
from pydantic import BaseModel
from typing import List, Dict, Any

app = FastAPI(title="Document Intelligence Service", version="2.0.0")

class SheetData(BaseModel):
    name: str
    is_hidden: bool
    data: List[List[str]]

class DocumentAnalysisResponse(BaseModel):
    filename: str
    sheets: List[SheetData]
    hidden_sheets: List[str]
    merged_cells: List[str]
    has_macros: bool
    has_comments: bool
    metadata: Dict[str, Any]

@app.post("/api/docintel/analyze", response_model=DocumentAnalysisResponse, status_code=status.HTTP_200_OK)
async def analyze_excel_document(file: UploadFile):
    """
    Performs deep structural analysis on Excel worksheets to detect hidden cells,
    macros, and embedded comments before forwarding contents to LLMs.
    """
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported format. Only openpyxl-compatible (.xlsx, .xlsm) files are accepted."
        )

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        result = DocumentAnalysisResponse(
            filename=file.filename,
            sheets=[],
            hidden_sheets=[],
            merged_cells=[],
            has_macros=wb.vba_code is not None,
            has_comments=False,
            metadata={}
        )

        for name in wb.sheetnames:
            sheet = wb[name]
            is_hidden = sheet.sheet_state == 'hidden'
            if is_hidden:
                result.hidden_sheets.append(name)

            grid_data = []
            for row in sheet.iter_rows(values_only=True):
                grid_data.append([str(cell) if cell is not None else '' for cell in row])

            # Detect comments in the sheet
            if hasattr(sheet, '_comments') and sheet._comments:
                result.has_comments = True

            for merged_range in sheet.merged_cells.ranges:
                result.merged_cells.append(str(merged_range))

            result.sheets.append(
                SheetData(name=name, is_hidden=is_hidden, data=grid_data)
            )

        return result
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Parsing error: {str(e)}")
